from core.parsing import parse_all_wallet_data
from core.datawork import build
from core.const import chains_list
from openpyxl import Workbook
from core.const import alphabet
from core.workpyxl import write_lists, write_data
wb = Workbook()
ws = wb.active

shapka = ['wallet','all','all farmed']
for i in chains_list:
    shapka.append(f"{i} coins")
    shapka.append(f"{i} farm")
    shapka.append(f"{i} big farm")

for i in range(len(shapka)):
    ws[f"{alphabet[i]}1"] = shapka[i]

with open('evm.txt') as file:
    wallets = file.readlines()
    wallets = [i.replace('\n','') for i in wallets]




for i,v in enumerate(wallets):
    try:
        data = parse_all_wallet_data(v)
        builded = build(data,v)
        row = i + 2
        write_lists(ws,builded[1],row)
        write_data(ws,builded[0],row)
        print(i+1,' done')
    except:
        print(f'writing error {i} {v}')
        print(data)
row = i + 3
for i in alphabet[1:]:
    ws[f"{i}{row}"] = f"=SUM({i}2:{i}{row-1})"
wb.save('evm.xlsx')
