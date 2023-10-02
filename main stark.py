from core.stark_datawork import stark_build
from core.stark_parsing import parse_all_wallet_data
from openpyxl import Workbook
from core.const import alphabet
from core.workpyxl import write_lists, write_data
wb = Workbook()
ws = wb.active

shapka = ['wallet','all','all farmed','big project','coins','big coin']

for i in range(len(shapka)):
    ws[f"{alphabet[i]}1"] = shapka[i]

with open('stark.txt') as file:
    wallets = file.readlines()
    wallets = [i.replace('\n','').replace(' ','') for i in wallets]


for i,v in enumerate(wallets):
    try:
        data = parse_all_wallet_data(v)
        builded = stark_build(data,v)
        row = i + 2
        write_lists(ws,builded[1],row)
        write_data(ws,builded[0],row)
        print(i+1,' done')
    except:
        print(f'writing error {i} {v}')
row = i + 3
for i in alphabet[1:5]:
    ws[f"{i}{row}"] = f"=SUM({i}2:{i}{row-1})"
wb.save('stark.xlsx')